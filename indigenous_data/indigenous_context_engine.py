#!/usr/bin/env python3
"""
Indigenous Context Engine for the VINCULUM Lesson Digester
==========================================================
Reads the Indigenous Math Data Repository (.xlsx) and matches
lesson topics to culturally relevant stories, themes, and problem contexts.

Usage:
    python indigenous_context_engine.py --grade 4 --domain "Geometry" --topic "symmetry"
    python indigenous_context_engine.py --grade 7 --domain "Ratios" --topic "proportional reasoning"
    python indigenous_context_engine.py --grade 2 --domain "Measurement" --topic "non-standard units"
    python indigenous_context_engine.py --test  (runs all built-in test cases)

Output: JSON with matched Indigenous contexts, sample problems, and Digester tags.
"""

import argparse
import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl


# ── Data Loading ──

def load_repository(xlsx_path):
    """Load all sheets from the Indigenous Math Data Repository."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    repo = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
        repo[sheet_name] = rows
    return repo


# ── Matching Engine ──

GRADE_BAND_MAP = {
    "K": "K-2", "0": "K-2", "1": "K-2", "2": "K-2",
    "3": "3-5", "4": "3-5", "5": "3-5",
    "6": "6-8", "7": "6-8", "8": "6-8",
    "9": "9-12", "10": "9-12", "11": "9-12", "12": "9-12",
}

DOMAIN_ALIASES = {
    "number": "Number & Operations",
    "operations": "Number & Operations",
    "counting": "Number & Operations",
    "place value": "Number & Operations",
    "addition": "Number & Operations",
    "subtraction": "Number & Operations",
    "multiplication": "Number & Operations",
    "division": "Number & Operations",
    "fractions": "Number & Operations",
    "measurement": "Measurement & Data",
    "data": "Measurement & Data",
    "statistics": "Statistics & Data",
    "probability": "Statistics & Probability",
    "geometry": "Geometry",
    "shapes": "Geometry",
    "symmetry": "Geometry",
    "area": "Geometry",
    "volume": "Geometry",
    "perimeter": "Geometry",
    "algebra": "Algebra & Functions",
    "functions": "Algebra & Functions",
    "equations": "Algebra & Functions",
    "patterns": "Algebra & Functions",
    "ratios": "Ratios & Proportions",
    "proportions": "Ratios & Proportions",
    "proportional": "Ratios & Proportions",
}


def normalize_domain(domain_input):
    """Map user input to standard domain names used in the repository."""
    d = domain_input.strip().lower()
    if d in DOMAIN_ALIASES:
        return DOMAIN_ALIASES[d]
    # Partial match
    for alias, canonical in DOMAIN_ALIASES.items():
        if alias in d or d in alias:
            return canonical
    return domain_input  # Return as-is if no match


def get_grade_band(grade_input):
    """Convert a grade number/string to a grade band."""
    g = str(grade_input).strip().upper().replace("GRADE ", "").replace("GR ", "")
    return GRADE_BAND_MAP.get(g, g)


def score_match(entry, grade_band, domain, topic_keywords):
    """Score how well a repository entry matches the query. Higher = better match."""
    score = 0

    # Grade band match
    entry_band = entry.get("Grade_Band") or entry.get("Grade_Range") or ""
    if grade_band in entry_band or entry_band in grade_band:
        score += 10
    elif any(g in entry_band for g in grade_band.split("-")):
        score += 5  # Partial overlap

    # Domain match
    entry_domain = entry.get("Math_Domain") or entry.get("Math_Domain") or ""
    entry_concepts = entry.get("Math_Concepts") or entry.get("Standard_Curriculum_Connection") or ""
    normalized = normalize_domain(domain)

    if normalized.lower() in entry_domain.lower():
        score += 10
    elif domain.lower() in entry_domain.lower() or domain.lower() in entry_concepts.lower():
        score += 7

    # Topic keyword matching
    searchable = " ".join([
        str(entry.get("Theme", "")),
        str(entry.get("Math_Concepts", "")),
        str(entry.get("Indigenous_Concept", "")),
        str(entry.get("Standard_Curriculum_Connection", "")),
        str(entry.get("Sample_Problem", "")),
        str(entry.get("Cultural_Context", "")),
        str(entry.get("Notes", "")),
    ]).lower()

    for kw in topic_keywords:
        if kw.lower() in searchable:
            score += 5

    return score


def match_contexts(repo, grade, domain, topic):
    """Find the best matching Indigenous contexts for a lesson."""
    grade_band = get_grade_band(grade)
    normalized_domain = normalize_domain(domain)
    topic_keywords = [w.strip() for w in topic.replace(",", " ").split() if len(w.strip()) > 2]

    # Score stories
    story_matches = []
    for entry in repo.get("Story_Theme_Repository", []):
        s = score_match(entry, grade_band, domain, topic_keywords)
        if s > 5:
            story_matches.append((s, entry))
    story_matches.sort(key=lambda x: -x[0])

    # Score crosswalk entries
    crosswalk_matches = []
    for entry in repo.get("Math_Domain_Crosswalk", []):
        s = score_match(entry, grade_band, domain, topic_keywords)
        if s > 5:
            crosswalk_matches.append((s, entry))
    crosswalk_matches.sort(key=lambda x: -x[0])

    # Find relevant resources
    resource_matches = []
    for entry in repo.get("Resource_Registry", []):
        grade_range = str(entry.get("Grade_Range", ""))
        math_domains = str(entry.get("Math_Domains", "")).lower()
        if (grade_band.split("-")[0] in grade_range or
            normalized_domain.lower() in math_domains or
            "all" in math_domains.lower()):
            resource_matches.append(entry)

    return {
        "query": {
            "grade": str(grade),
            "grade_band": grade_band,
            "domain": domain,
            "normalized_domain": normalized_domain,
            "topic": topic,
        },
        "stories": [
            {
                "match_score": s,
                "id": e.get("ID"),
                "theme": e.get("Theme"),
                "cultural_context": e.get("Cultural_Context"),
                "nations": e.get("Nation(s)"),
                "math_concepts": e.get("Math_Concepts"),
                "cra_phase": e.get("CRA_Phase"),
                "sample_problem": e.get("Sample_Problem"),
                "mn_aligned": e.get("MN_Benchmark_Aligned"),
                "notes": e.get("Notes"),
            }
            for s, e in story_matches[:5]
        ],
        "crosswalk_tags": [
            {
                "match_score": s,
                "digester_tag": e.get("Digester_Tag"),
                "indigenous_concept": e.get("Indigenous_Concept"),
                "nations": e.get("Nation(s)"),
                "curriculum_connection": e.get("Standard_Curriculum_Connection"),
                "cra_phase": e.get("CRA_Phase_Best_Fit"),
            }
            for s, e in crosswalk_matches[:5]
        ],
        "resources": [
            {
                "name": e.get("Resource_Name"),
                "type": e.get("Type"),
                "evidence_level": e.get("Evidence_Level"),
                "integration_notes": e.get("Digester_Integration_Notes"),
            }
            for e in resource_matches[:5]
        ],
        "integration_guidance": build_guidance(story_matches, crosswalk_matches, grade_band),
    }


def build_guidance(stories, crosswalk, grade_band):
    """Generate plain-language integration guidance for the Digester."""
    if not stories and not crosswalk:
        return {
            "status": "NO_MATCH",
            "message": "No Indigenous context matches found for this lesson. Consider broadening the topic keywords or checking the domain mapping.",
        }

    guidance = {
        "status": "MATCHES_FOUND",
        "total_story_matches": len(stories),
        "total_crosswalk_matches": len(crosswalk),
        "recommendations": [],
    }

    if stories:
        best = stories[0][1]
        guidance["primary_recommendation"] = {
            "theme": best.get("Theme"),
            "action": f"Replace generic problem context with: {best.get('Cultural_Context')}",
            "sample_problem": best.get("Sample_Problem"),
            "cra_phase": best.get("CRA_Phase"),
        }
        guidance["recommendations"].append(
            f"Use '{best.get('Theme')}' ({best.get('Nation(s)')}) as the story context for this lesson's problem set."
        )

    if crosswalk:
        best_cw = crosswalk[0][1]
        guidance["recommendations"].append(
            f"Map to Digester tag: {best_cw.get('Digester_Tag')} — {best_cw.get('Indigenous_Concept')}"
        )

    guidance["recommendations"].append(
        "IMPORTANT: All Indigenous content should be flagged for review by Indigenous educators before deployment."
    )
    guidance["recommendations"].append(
        "Use present tense for living cultures. Specify which nation(s) the content comes from."
    )

    return guidance


# ── Test Suite ──

TEST_CASES = [
    {"grade": "K", "domain": "counting", "topic": "counting number words Ojibwe classifying",
     "expect_theme": "Counting in Ojibwe"},
    {"grade": "2", "domain": "measurement", "topic": "non-standard units measuring body",
     "expect_theme": "Moccasin Making"},
    {"grade": "4", "domain": "geometry", "topic": "symmetry shapes reflection",
     "expect_theme": "Beadwork Design"},
    {"grade": "5", "domain": "geometry", "topic": "area perimeter rectangles",
     "expect_theme": "Fish Rack Building"},
    {"grade": "3", "domain": "data", "topic": "data collection who collected perspective",
     "expect_theme": "Data Sovereignty"},
    {"grade": "6", "domain": "ratios", "topic": "proportional reasoning rates",
     "expect_theme": "Salmon Drying Ratios"},
    {"grade": "7", "domain": "geometry", "topic": "3D surface area volume prisms",
     "expect_theme": "Smokehouse Geometry"},
    {"grade": "8", "domain": "statistics", "topic": "statistical analysis data comparison",
     "expect_theme": "Kayak Design"},
    {"grade": "4", "domain": "number", "topic": "multiplication division equivalence trading",
     "expect_theme": "Trading and Exchange"},
    {"grade": "10", "domain": "algebra", "topic": "exponential functions modeling population",
     "expect_theme": "Demographic Modeling"},
    {"grade": "11", "domain": "statistics", "topic": "regression prediction language speaker population",
     "expect_theme": "Language Revitalization"},
    {"grade": "9", "domain": "algebra", "topic": "linear programming optimization",
     "expect_theme": "Resource Management"},
]


def run_tests(repo):
    """Run all test cases and report results."""
    print("\n" + "=" * 70)
    print("INDIGENOUS CONTEXT ENGINE — TEST SUITE")
    print("=" * 70)

    passed = 0
    failed = 0

    for i, tc in enumerate(TEST_CASES, 1):
        result = match_contexts(repo, tc["grade"], tc["domain"], tc["topic"])
        top_theme = result["stories"][0]["theme"] if result["stories"] else "NONE"
        expected = tc["expect_theme"]
        match = expected.lower() in top_theme.lower() or top_theme.lower() in expected.lower()

        status = "PASS" if match else "FAIL"
        if match:
            passed += 1
        else:
            failed += 1

        print(f"\n  Test {i:2d} [{status}] Grade {tc['grade']:>2s} | {tc['domain']:<14s} | '{tc['topic'][:35]}...'")
        print(f"          Expected: {expected}")
        print(f"          Got:      {top_theme} (score: {result['stories'][0]['match_score'] if result['stories'] else 0})")

        if result["crosswalk_tags"]:
            print(f"          Tag:      {result['crosswalk_tags'][0]['digester_tag']}")

        if not match and result["stories"]:
            print(f"          Alt matches: {[s['theme'] for s in result['stories'][:3]]}")

    print(f"\n{'=' * 70}")
    print(f"  RESULTS: {passed}/{passed+failed} passed ({passed/(passed+failed)*100:.0f}%)")
    if failed:
        print(f"  {failed} tests need tuning — check domain aliases or topic keywords")
    print(f"{'=' * 70}\n")
    return failed == 0


# ── Demo Mode ──

def run_demo(repo):
    """Run a few demo queries to show how the engine works."""
    print("\n" + "=" * 70)
    print("INDIGENOUS CONTEXT ENGINE — DEMO")
    print("=" * 70)

    demos = [
        ("4", "geometry", "symmetry reflection shapes"),
        ("6", "ratios", "proportional reasoning unit rates"),
        ("2", "measurement", "non-standard units body measuring"),
        ("10", "statistics", "regression data analysis prediction"),
    ]

    for grade, domain, topic in demos:
        result = match_contexts(repo, grade, domain, topic)
        print(f"\n{'─' * 60}")
        print(f"  QUERY: Grade {grade} | {domain} | '{topic}'")
        print(f"{'─' * 60}")

        if result["stories"]:
            s = result["stories"][0]
            print(f"\n  TOP MATCH: {s['theme']}")
            print(f"  Nations:   {s['nations']}")
            print(f"  Context:   {s['cultural_context']}")
            print(f"  CRA Phase: {s['cra_phase']}")
            print(f"  MN Aligned: {s['mn_aligned']}")
            print(f"\n  SAMPLE PROBLEM:")
            # Word wrap the sample problem
            prob = s['sample_problem'] or ""
            words = prob.split()
            line = "    "
            for w in words:
                if len(line) + len(w) > 68:
                    print(line)
                    line = "    " + w + " "
                else:
                    line += w + " "
            if line.strip():
                print(line)

        if result["crosswalk_tags"]:
            t = result["crosswalk_tags"][0]
            print(f"\n  DIGESTER TAG: {t['digester_tag']}")
            print(f"  Concept:      {t['indigenous_concept']}")

        g = result["integration_guidance"]
        if g.get("recommendations"):
            print(f"\n  GUIDANCE:")
            for rec in g["recommendations"][:2]:
                print(f"    → {rec[:75]}")

    print(f"\n{'=' * 70}\n")


# ── Main ──

def main():
    parser = argparse.ArgumentParser(
        description="Indigenous Context Engine for the VINCULUM Lesson Digester",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python indigenous_context_engine.py --grade 4 --domain geometry --topic "symmetry reflection"
  python indigenous_context_engine.py --grade 7 --domain ratios --topic "proportional reasoning"
  python indigenous_context_engine.py --test
  python indigenous_context_engine.py --demo
  python indigenous_context_engine.py --grade 2 --domain measurement --topic "non-standard units" --json
        """,
    )
    parser.add_argument("--grade", type=str, help="Grade level (K, 1-12)")
    parser.add_argument("--domain", type=str, help="Math domain (geometry, ratios, measurement, etc.)")
    parser.add_argument("--topic", type=str, help="Lesson topic keywords")
    parser.add_argument("--test", action="store_true", help="Run the test suite")
    parser.add_argument("--demo", action="store_true", help="Run demo queries")
    parser.add_argument("--json", action="store_true", help="Output raw JSON (for Digester ingestion)")
    parser.add_argument("--xlsx", type=str, default=None, help="Path to the Indigenous Math Data Repository xlsx")

    args = parser.parse_args()

    # Find the xlsx file
    xlsx_path = args.xlsx
    if not xlsx_path:
        candidates = [
            "Indigenous_Math_Data_Repository.xlsx",
            os.path.join(os.path.dirname(__file__), "Indigenous_Math_Data_Repository.xlsx"),
        ]
        for c in candidates:
            if os.path.exists(c):
                xlsx_path = c
                break

    if not xlsx_path or not os.path.exists(xlsx_path):
        print(f"ERROR: Cannot find Indigenous_Math_Data_Repository.xlsx")
        print(f"  Use --xlsx <path> to specify location")
        sys.exit(1)

    repo = load_repository(xlsx_path)
    print(f"Loaded repository: {sum(len(v) for v in repo.values())} total entries across {len(repo)} sheets")

    if args.test:
        success = run_tests(repo)
        sys.exit(0 if success else 1)

    if args.demo:
        run_demo(repo)
        sys.exit(0)

    if not all([args.grade, args.domain, args.topic]):
        parser.print_help()
        print("\n  TIP: Try --demo or --test first to see how it works!")
        sys.exit(1)

    result = match_contexts(repo, args.grade, args.domain, args.topic)

    if args.json:
        print(json.dumps(result, indent=2, default=str))
    else:
        # Human-readable output
        print(f"\n{'─' * 60}")
        print(f"  Grade {args.grade} | {args.domain} | '{args.topic}'")
        print(f"{'─' * 60}")

        if result["stories"]:
            for i, s in enumerate(result["stories"][:3], 1):
                print(f"\n  Match #{i} (score {s['match_score']}): {s['theme']}")
                print(f"    Nations:  {s['nations']}")
                print(f"    Context:  {s['cultural_context']}")
                print(f"    CRA:      {s['cra_phase']}")
                print(f"    Problem:  {(s['sample_problem'] or '')[:80]}...")
        else:
            print("\n  No story matches found.")

        if result["crosswalk_tags"]:
            print(f"\n  Digester Tags:")
            for t in result["crosswalk_tags"][:3]:
                print(f"    [{t['digester_tag']}] {t['indigenous_concept']}")

        g = result["integration_guidance"]
        if g.get("recommendations"):
            print(f"\n  Integration Guidance:")
            for rec in g["recommendations"]:
                print(f"    → {rec}")

        print()


if __name__ == "__main__":
    main()
